"""
TPM × FW Rate Card Compliance Validator (FabSol) — v2.6
Date: 2026-06-29

v2.5 changes:
✓ FIXED: Compliance sheet no longer shows ANY metadata/filter banner
         (just column headers + data, autofilter ready)
✓ MOVED: ALL filter info → hidden Metadata sheet
         (CD_Category, Quarter_Year, TPM Source, Rate Card Source, Tolerance,
          Version, Generated, Working Dir, Total Rows)

v2.4 features preserved:
✓ Terminal Justification Breakdown w/ colored bars
✓ Summary sheet — 3-color scale + data bars
✓ CD_Category == 'FAB SOL' + Quarter_Year hard filters
✓ TPM sheet picker + Rich progress bars + emoji status lines
"""
from __future__ import annotations

import os
import re
import sys
import warnings
import logging
import contextlib
import io as _io
import traceback
from datetime import date, datetime
from pathlib import Path

import polars as pl
import pandas as pd
import questionary
from questionary import Style as QStyle
from openpyxl import load_workbook
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import (
    Progress, SpinnerColumn, TextColumn, BarColumn,
    TimeElapsedColumn, MofNCompleteColumn,
)
from rich import box

# ============================================================
# VERSION
# ============================================================
__version__       = "2.6"
__version_date__  = "2026-06-29"
__version_notes__ = "Reordered pickers (files first, then sheets) + 'FW' text removed"

console = Console()

if os.name == "nt":
    try:
        os.system(f"title TPM FabSol Compliance v{__version__}")
    except Exception:
        pass

warnings.filterwarnings("ignore")
logging.getLogger("polars").setLevel(logging.ERROR)

# ============================================================
# PATH
# ============================================================
if getattr(sys, "frozen", False):
    WORK_DIR = Path(sys.executable).parent
else:
    WORK_DIR = Path(__file__).parent

# ============================================================
# CONFIG
# ============================================================
CD_CATEGORY_KEEP = "FAB SOL"
TOLERANCE = 0.5

DROP_COLUMNS = ["RSP CCBT", "GAP", "promo mechanic"]
COLUMN_RENAME_MAP = {"WPRM": "RSP Promo"}

REQUIRED_COLUMNS = [
    "PromoGroupProductDesc",
    "Instore_Start",
    "Instore_End",
    "TPM_InvestmentDescription",
    "RSP Promo",
]


def col_letter_to_idx(letter: str) -> int:
    letter = letter.upper().strip()
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - ord('A') + 1)
    return n - 1


PERIOD_COL_MAP = {
    "Weekly":     [col_letter_to_idx(c) for c in ("AJ", "L", "V", "X")],
    "Bi-weekly":  [col_letter_to_idx(c) for c in ("J", "N")],
    "Tri-weekly": [col_letter_to_idx(c) for c in ("J", "N")],
    "Monthly":    [col_letter_to_idx(c) for c in ("H", "AA")],
}

PERIOD_BUCKETS = {"Weekly": 7, "Bi-weekly": 14, "Tri-weekly": 21, "Monthly": 30}

COLOR_MAP_HEX = {
    "Comply":                 "#C6EFCE",
    "Not Comply":             "#FFC7CE",
    "NO rate card available": "#FFEB9C",
    "Missing Final_Rsp":      "#D9D9D9",
}

JUST_COLOR = {
    "Comply":                 "green",
    "Not Comply":             "red",
    "NO rate card available": "yellow",
    "Missing Final_Rsp":      "white",
    "Quarter mismatch":       "dark_orange",
}
JUST_BAR_COLOR = {
    "Comply":                 "green",
    "Not Comply":             "red",
    "NO rate card available": "yellow",
    "Missing Final_Rsp":      "grey50",
    "Quarter mismatch":       "dark_orange",
}

QSTYLE = QStyle([
    ("qmark",       "fg:#00aaff bold"),
    ("question",    "bold"),
    ("answer",      "fg:#00aaff bold"),
    ("pointer",     "fg:#ff8800 bold"),
    ("highlighted", "fg:#00aaff bold noreverse"),
    ("selected",    "noinherit"),
    ("instruction", "fg:#888888 italic"),
])

# ============================================================
# UI HELPERS
# ============================================================
def banner():
    console.print()
    console.print(Panel.fit(
        f"[bold cyan]TPM × FW Rate Card Compliance Validator (FabSol)[/]\n"
        f"[bold]Version {__version__}[/]  [dim]({__version_date__})[/]\n"
        f"[dim italic]{__version_notes__}[/]\n"
        f"[dim]Working dir: {WORK_DIR}[/]",
        border_style="cyan",
        box=box.DOUBLE,
        title=f"[bold magenta]v{__version__}[/]",
        subtitle=f"[dim]Python {sys.version_info.major}.{sys.version_info.minor}[/]",
    ))
    console.print()


def list_xlsx_files(folder: Path) -> list[Path]:
    return sorted([
        p for p in folder.iterdir()
        if p.suffix.lower() in (".xlsx", ".xlsb", ".xls")
        and not p.name.startswith("~$")
        and not p.name.startswith("FabSol_Compliance_")
        and not p.name.startswith("TPM_FabSol_Compliance_")
        and not p.name.startswith("TPM_Rate_Card_Compliance_Output")
        and p.is_file()
    ], key=lambda p: p.stat().st_mtime, reverse=True)


def pick_file(files: list[Path], prompt: str) -> Path:
    choices = [questionary.Choice(title=f.name, value=f) for f in files]
    answer = questionary.select(
        prompt, choices=choices, style=QSTYLE,
        instruction="(↑/↓ to move, Enter to select, Esc to cancel)",
    ).ask()
    if answer is None:
        console.print("[yellow]Cancelled by user[/]")
        sys.exit(0)
    return answer


def pick_sheet(prompt: str, sheets: list[str]) -> str:
    answer = questionary.select(
        prompt, choices=sheets, style=QSTYLE,
        instruction="(↑/↓ to move, Enter to select, Esc to cancel)",
    ).ask()
    if answer is None:
        console.print("[yellow]Cancelled by user[/]")
        sys.exit(0)
    return answer


def get_sheet_names(path: Path) -> list[str]:
    if path.suffix.lower() == ".xlsb":
        from pyxlsb import open_workbook
        with open_workbook(path) as wb:
            return wb.sheets
    wb = load_workbook(path, read_only=True, data_only=False)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def derive_target_quarter(sheet_name: str) -> str | None:
    m = re.search(r"Q\s*(\d).*?(\d{4})", sheet_name)
    return f"Q{m.group(1)} {m.group(2)}" if m else None


def make_progress() -> Progress:
    return Progress(
        SpinnerColumn(style="cyan"),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(bar_width=40),
        MofNCompleteColumn(),
        TextColumn("[dim]•[/]"),
        TimeElapsedColumn(),
        console=console,
        transient=False,
    )

# ============================================================
# POLARS PROCESSING
# ============================================================
def _columns_to_drop(df_columns: list) -> list:
    drop_lower = {c.lower() for c in DROP_COLUMNS}
    result = []
    for col in df_columns:
        c = str(col).strip()
        if c.lower() in drop_lower:
            result.append(col)
        elif c.lower().startswith("unnamed") or c.startswith("__UNNAMED__"):
            result.append(col)
    return result


def _apply_column_renames(df: pl.DataFrame) -> pl.DataFrame:
    cols = set(df.columns)
    for original, standard in COLUMN_RENAME_MAP.items():
        if original in cols and standard not in cols:
            df = df.rename({original: standard})
            console.print(f"   🔁 Renamed [yellow]{original}[/] → [cyan]{standard}[/]")
    return df


def load_tpm(tpm_path: Path, sheet_name: str, progress: Progress) -> pl.DataFrame:
    task = progress.add_task("Loading TPM file (calamine)", total=1)
    if tpm_path.suffix.lower() == ".xlsb":
        df_pd = pd.read_excel(tpm_path, sheet_name=sheet_name, engine="pyxlsb")
        df = pl.from_pandas(df_pd)
    else:
        with contextlib.redirect_stderr(_io.StringIO()):
            df = pl.read_excel(tpm_path, sheet_name=sheet_name, engine="calamine")
    progress.update(task, advance=1)
    return df


def validate_and_filter(df: pl.DataFrame, progress: Progress) -> pl.DataFrame:
    task = progress.add_task("Validating & filtering", total=2)

    df = _apply_column_renames(df)
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        hint = ""
        if "RSP Promo" in missing and "WPRM" in df.columns:
            hint = "\n[Hint] 'WPRM' present but not renamed."
        console.print(f"[red]❌ Missing required columns:[/] {missing}{hint}")
        sys.exit(1)
    progress.update(task, advance=1)

    if "CD_Category" not in df.columns:
        console.print("[red]❌ Column 'CD_Category' not found in TPM[/]")
        sys.exit(1)
    before = df.height
    df = df.filter(
        pl.col("CD_Category").cast(pl.Utf8).str.strip_chars() == CD_CATEGORY_KEEP
    )
    console.print(f"   🎯 CD_Category=='{CD_CATEGORY_KEEP}': kept {df.height:,} / {before:,}")
    progress.update(task, advance=1)
    return df


def _extract_promo_type(text):
    if not isinstance(text, str): return None
    t = text.upper()
    if "BOGO"   in t: return "BOGO"
    if "2F1"    in t: return "2F1"
    if "LAKSUE" in t: return "LAKSUE"
    if "_PO"    in t: return "PO"
    return None


def _quarter_year(dt):
    if dt is None: return None
    try:
        return f"Q{((dt.month - 1) // 3) + 1} {dt.year}"
    except Exception:
        return None


def _closest_period(days):
    if days is None: return None
    try: d = int(days)
    except (TypeError, ValueError): return None
    return min(PERIOD_BUCKETS, key=lambda k: abs(PERIOD_BUCKETS[k] - d))


def add_helper_columns(df: pl.DataFrame, target_quarter: str | None,
                       progress: Progress) -> pl.DataFrame:
    task = progress.add_task("Building helper columns", total=8)

    df = df.with_columns(
        pl.col("TPM_InvestmentDescription")
          .map_elements(_extract_promo_type, return_dtype=pl.Utf8)
          .alias("Promo_Type")
    ); progress.update(task, advance=1)

    df = df.with_columns(
        pl.col("RSP Promo").cast(pl.Float64, strict=False).alias("Final_Rsp")
    ); progress.update(task, advance=1)

    df = df.with_columns(
        (pl.col("Instore_End").cast(pl.Date, strict=False)
         - pl.col("Instore_Start").cast(pl.Date, strict=False))
        .dt.total_days().alias("Duration_Days")
    ); progress.update(task, advance=1)

    df = df.with_columns(
        pl.col("Duration_Days")
          .map_elements(_closest_period, return_dtype=pl.Utf8)
          .alias("Period")
    ); progress.update(task, advance=1)

    df = df.with_columns(
        pl.col("Instore_Start").cast(pl.Date, strict=False)
          .map_elements(_quarter_year, return_dtype=pl.Utf8)
          .alias("Quarter_Year")
    ); progress.update(task, advance=1)

    if target_quarter:
        before = df.height
        df = df.filter(pl.col("Quarter_Year") == target_quarter)
        console.print(f"   📅 Quarter_Year=='{target_quarter}': kept {df.height:,} / {before:,}")
    else:
        console.print("[yellow]⚠ Could not derive Q_Y from sheet name — skipping Quarter filter[/]")
    progress.update(task, advance=1)

    pt_counts = {row["Promo_Type"]: row["len"]
                 for row in df.group_by("Promo_Type").len().iter_rows(named=True)}
    console.print(f"   Promo_Type counts: {pt_counts}")
    progress.update(task, advance=1)

    pd_counts = {row["Period"]: row["len"]
                 for row in df.group_by("Period").len().iter_rows(named=True)}
    console.print(f"   Period counts:     {pd_counts}")
    progress.update(task, advance=1)
    return df

# ============================================================
# RATE CARD INDEXING
# ============================================================
def parse_set_cell(v) -> list:
    if v is None or (isinstance(v, float) and pd.isna(v)): return []
    if isinstance(v, (int, float)):
        return [round(float(v), 1)]
    s = str(v).strip()
    if not s or s.lower() == "nan": return []
    parts = [p.strip() for p in re.split(r"[/,]", s) if p.strip()]
    out = []
    for p in parts:
        try:
            out.append(round(float(p), 1))
        except ValueError:
            try:
                out.append(round(float(p.rstrip("%").strip()), 1))
            except ValueError:
                out.append(p.upper())
    return out


def format_set(values: list) -> str | None:
    if not values: return None
    parts = []
    for v in values:
        parts.append(v if isinstance(v, str) else f"{v:g}")
    return "/".join(parts)


def find_header_row(ws, keyword="Promo Group"):
    for r in range(1, 30):
        for c in range(1, 30):
            v = ws.cell(r, c).value
            if isinstance(v, str) and keyword.lower() in v.lower():
                return r
    raise RuntimeError(f"Header row containing '{keyword}' not found")


def build_rate_card_index(path: Path, sheet: str, progress: Progress):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    header_row = find_header_row(ws, "Promo Group")
    console.print(f"   Header row at: {header_row}")

    ppg_col_idx = None
    period_col_idx = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str):
            vl = v.strip().lower()
            if "promo group" in vl or vl == "ppg":
                ppg_col_idx = c
            elif vl == "period":
                period_col_idx = c
    if ppg_col_idx is None:
        raise RuntimeError("PPG column not found in Rate Card header")

    index: dict[str, dict[str, list]] = {}
    first_data_row = header_row + 1
    task = progress.add_task("Indexing Rate Card", total=ws.max_row - header_row)
    row_count = 0
    for r in range(first_data_row, ws.max_row + 1):
        ppg = ws.cell(r, ppg_col_idx).value
        if ppg is None or str(ppg).strip() == "":
            progress.update(task, advance=1); continue
        key = str(ppg).strip()
        period_val = ws.cell(r, period_col_idx).value if period_col_idx else None
        period_key = str(period_val).strip() if period_val else None

        per_period: dict[str, list] = index.setdefault(key, {})
        for period_name, col_idx_list in PERIOD_COL_MAP.items():
            if period_key and period_key != period_name:
                continue
            collected: list = []
            for ci in col_idx_list:
                cell_val = ws.cell(r, ci + 1).value
                collected.extend(parse_set_cell(cell_val))
            if collected:
                per_period.setdefault(period_name, []).extend(collected)
        row_count += 1
        progress.update(task, advance=1)

    console.print(f"   Indexed [cyan]{row_count}[/] PPG rows")
    wb.close()
    return index


def lookup_rate_number(index, ppg, period):
    if ppg is None or period is None: return None, []
    key = str(ppg).strip()
    bucket = index.get(key)
    if not bucket: return None, []
    vals = bucket.get(period, [])
    return (format_set(vals) if vals else None), vals

# ============================================================
# COMPLIANCE EVALUATION
# ============================================================
def justify(rate_str, final_rsp, vals):
    if not rate_str:
        return "NO rate card available"
    if final_rsp is None:
        return "Missing Final_Rsp"
    for v in vals:
        try:
            if abs(float(v) - float(final_rsp)) <= TOLERANCE:
                return "Comply"
        except (TypeError, ValueError):
            continue
    return "Not Comply"


def evaluate(df: pl.DataFrame, rc_index, progress: Progress) -> pl.DataFrame:
    task = progress.add_task("Evaluating compliance", total=df.height)
    rate_nums: list[str | None] = []
    justifications: list[str] = []
    for row in df.iter_rows(named=True):
        rs, vals = lookup_rate_number(
            rc_index, row.get("PromoGroupProductDesc"), row.get("Period")
        )
        j = justify(rs, row.get("Final_Rsp"), vals)
        rate_nums.append(rs)
        justifications.append(j)
        progress.update(task, advance=1)

    df = df.with_columns([
        pl.Series("Rate_Number", rate_nums, dtype=pl.Utf8),
        pl.Series("Justification", justifications, dtype=pl.Utf8),
    ])
    return df

# ============================================================
# SUMMARY BUILDER
# ============================================================
JUST_ORDER = ["Comply", "Not Comply", "NO rate card available", "Missing Final_Rsp"]


def build_summary(df: pl.DataFrame) -> pl.DataFrame:
    grp_cols = [c for c in ["Customer", "Brand", "Period"] if c in df.columns]
    if not grp_cols:
        grp_cols = ["Period"] if "Period" in df.columns else []

    if not grp_cols:
        agg = df.group_by("Justification").len().rename({"len": "Count"})
        return agg.sort("Justification")

    pivot = (
        df.group_by(grp_cols + ["Justification"]).len()
          .rename({"len": "Count"})
          .pivot(values="Count", index=grp_cols, on="Justification",
                 aggregate_function="first")
          .fill_null(0)
    )

    for j in JUST_ORDER:
        if j not in pivot.columns:
            pivot = pivot.with_columns(pl.lit(0).alias(j))

    pivot = pivot.with_columns(
        sum(pl.col(j) for j in JUST_ORDER).alias("Total")
    ).with_columns(
        (pl.col("Comply") / pl.when(pl.col("Total") == 0).then(1).otherwise(pl.col("Total")))
        .alias("Compliance_Rate")
    )

    ordered = grp_cols + JUST_ORDER + ["Total", "Compliance_Rate"]
    pivot = pivot.select([c for c in ordered if c in pivot.columns])
    return pivot.sort("Total", descending=True)

# ============================================================
# OUTPUT — xlsxwriter
# Sheet order: Compliance → Summary → Metadata (LAST, hidden)
# ============================================================
def write_output(df: pl.DataFrame, summary: pl.DataFrame, out_path: Path,
                 tpm_path: Path, tpm_sheet: str,
                 rate_path: Path, rate_sheet: str,
                 target_quarter: str | None,
                 progress: Progress):
    import xlsxwriter

    final_drop = _columns_to_drop(df.columns)
    if final_drop:
        df = df.drop(final_drop)

    task = progress.add_task("Writing Excel (xlsxwriter)", total=df.height)
    wb = xlsxwriter.Workbook(out_path, {"constant_memory": False})

    # ---------- Formats ----------
    title_fmt = wb.add_format({
        "bold": True, "font_size": 14, "font_color": "white",
        "bg_color": "#1F4E78", "align": "left", "valign": "vcenter",
        "border": 1,
    })
    meta_label_fmt = wb.add_format({
        "bold": True, "font_color": "#1F4E78", "bg_color": "#DDEBF7",
        "align": "left", "valign": "vcenter", "border": 1,
    })
    meta_value_fmt = wb.add_format({
        "bg_color": "#F2F2F2", "align": "left",
        "valign": "vcenter", "border": 1,
    })
    header_fmt = wb.add_format({
        "bold": True, "bg_color": "#305496", "font_color": "white",
        "border": 1, "align": "center", "valign": "vcenter",
    })
    fmt_map = {k: wb.add_format({"bg_color": v, "border": 1})
               for k, v in COLOR_MAP_HEX.items()}

    # ============================================================
    # SHEET 1 — Compliance (CLEAN: column headers + data only)
    # ============================================================
    ws = wb.add_worksheet("Compliance")
    cols = df.columns

    # Row 0 = column headers, data starts row 1
    for ci, col in enumerate(cols):
        ws.write(0, ci, col, header_fmt)
    ws.autofilter(0, 0, df.height, len(cols) - 1)

    try:
        just_idx = cols.index("Justification")
    except ValueError:
        just_idx = -1

    for ri, row in enumerate(df.iter_rows(), start=1):
        for ci, val in enumerate(row):
            if ci == just_idx and val in fmt_map:
                ws.write(ri, ci, val, fmt_map[val])
            else:
                if val is None:
                    ws.write_blank(ri, ci, None)
                else:
                    ws.write(ri, ci, val)
        progress.update(task, advance=1)

    for ci, col in enumerate(cols):
        ws.set_column(ci, ci, max(12, min(36, len(str(col)) + 6)))

    # ============================================================
    # SHEET 2 — Summary (with color scale + data bars)
    # ============================================================
    ws_sum = wb.add_worksheet("Summary")
    ws_sum.set_row(0, 24)
    sum_cols = summary.columns
    n_sum = max(len(sum_cols), 4)
    ws_sum.merge_range(0, 0, 0, n_sum - 1,
                       f"Compliance Summary  —  CD_Category: {CD_CATEGORY_KEEP}  |  "
                       f"Quarter: {target_quarter or '—'}",
                       title_fmt)
    for ci, col in enumerate(sum_cols):
        ws_sum.write(2, ci, col, header_fmt)

    pct_fmt  = wb.add_format({"num_format": "0.0%", "border": 1})
    int_fmt  = wb.add_format({"num_format": "#,##0", "border": 1})
    text_fmt = wb.add_format({"border": 1})

    for ri, row in enumerate(summary.iter_rows(named=True), start=3):
        for ci, col in enumerate(sum_cols):
            val = row[col]
            if col == "Compliance_Rate":
                ws_sum.write(ri, ci, val if val is not None else 0, pct_fmt)
            elif col in JUST_ORDER or col == "Total":
                ws_sum.write(ri, ci, val if val is not None else 0, int_fmt)
            else:
                ws_sum.write(ri, ci, val if val is not None else "", text_fmt)

    if "Compliance_Rate" in sum_cols and summary.height > 0:
        cr_col_idx = sum_cols.index("Compliance_Rate")
        last_row = 2 + summary.height
        ws_sum.conditional_format(
            3, cr_col_idx, last_row, cr_col_idx,
            {
                "type": "3_color_scale",
                "min_type": "num",  "min_value": 0,   "min_color": "#F8696B",
                "mid_type": "num",  "mid_value": 0.5, "mid_color": "#FFEB84",
                "max_type": "num",  "max_value": 1,   "max_color": "#63BE7B",
            },
        )

    if "Total" in sum_cols and summary.height > 0:
        tot_col_idx = sum_cols.index("Total")
        last_row = 2 + summary.height
        ws_sum.conditional_format(
            3, tot_col_idx, last_row, tot_col_idx,
            {"type": "data_bar", "bar_color": "#5B9BD5", "bar_solid": True},
        )

    for ci, col in enumerate(sum_cols):
        ws_sum.set_column(ci, ci, max(14, min(28, len(str(col)) + 4)))

    # ============================================================
    # SHEET 3 (LAST, HIDDEN) — Metadata (holds ALL filter/source info)
    # ============================================================
    ws_meta = wb.add_worksheet("Metadata")
    ws_meta.set_row(0, 22)
    ws_meta.merge_range(0, 0, 0, 1, "Run Metadata (hidden)", title_fmt)
    meta_rows = [
        ("Version",                    f"v{__version__}  ({__version_date__})"),
        ("Generated",                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Working Dir",                str(WORK_DIR)),
        ("CD_Category Filter",         CD_CATEGORY_KEEP),
        ("Quarter_Year Filter",        target_quarter or "— (not applied)"),
        ("TPM Source",                 f"{tpm_path.name}  →  [{tpm_sheet}]"),
        ("Rate Card Source",           f"{rate_path.name}  →  [{rate_sheet}]"),
        ("Tolerance",                  f"± {TOLERANCE}"),
        ("Total Rows (after filters)", f"{df.height:,}"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws_meta.write(i, 0, k, meta_label_fmt)
        ws_meta.write(i, 1, str(v), meta_value_fmt)
    ws_meta.set_column(0, 0, 28)
    ws_meta.set_column(1, 1, 70)
    ws_meta.hide()

    wb.close()

# ============================================================
# RICH SUMMARY DISPLAY — v1.9 style with colored bars
# ============================================================
def display_summary(df: pl.DataFrame, out_path: Path):
    counts_df = df.group_by("Justification").len().sort("len", descending=True)
    counts = {row["Justification"]: row["len"] for row in counts_df.iter_rows(named=True)}
    total = df.height or 1
    max_count = max(counts.values()) if counts else 1
    BAR_WIDTH = 30

    table = Table(
        title=f"📊 [bold]Justification Breakdown[/]  [dim](v{__version__})[/]",
        box=box.ROUNDED, show_header=True, header_style="bold cyan",
        title_justify="center",
    )
    table.add_column("Category", style="bold", no_wrap=True)
    table.add_column("Count",    justify="right", style="cyan")
    table.add_column("%",        justify="right", style="cyan")
    table.add_column("Bar",      no_wrap=True)

    for cat, cnt in counts.items():
        pct = cnt / total
        bar_len = max(1, int(round(BAR_WIDTH * cnt / max_count)))
        bar_color = JUST_BAR_COLOR.get(cat, "cyan")
        cat_color = JUST_COLOR.get(cat, "white")
        bar = f"[{bar_color}]{'█' * bar_len}[/]"
        table.add_row(
            f"[{cat_color}]{cat}[/]",
            f"{cnt:,}",
            f"{pct:.1%}",
            bar,
        )

    table.add_section()
    table.add_row(
        "[bold]Total[/]",
        f"[bold]{df.height:,}[/]",
        "[bold]100.0%[/]",
        "",
    )

    console.print()
    console.print(table)
    console.print(f"\n[green bold]✅ Saved:[/] {out_path}\n")

# ============================================================
# MAIN
# ============================================================

# ============================================================
# MAIN
# ============================================================
def main():
    banner()
    console.print(f"[dim]Startup — FabSol Validator v{__version__}[/]")

    files = list_xlsx_files(WORK_DIR)
    if not files:
        console.print(f"[red]❌ No Excel files in {WORK_DIR}[/]")
        sys.exit(1)

    # 1) Pick TPM file
    tpm_path  = pick_file(files, "Select the TPM file:")

    # 2) Pick Rate Card file
    rate_path = pick_file(files, "Select the Rate Card file:")

    # 3) Pick TPM sheet
    tpm_sheets = get_sheet_names(tpm_path)
    tpm_sheet  = pick_sheet(f"Select sheet in TPM ({tpm_path.name}):", tpm_sheets)

    # 4) Pick Rate Card sheet
    rate_sheets = get_sheet_names(rate_path)
    rate_sheet  = pick_sheet(f"Select sheet in Rate Card ({rate_path.name}):", rate_sheets)

    target_quarter = derive_target_quarter(rate_sheet)
    if target_quarter:
        console.print(f"[dim]Rate Card target quarter: [cyan]{target_quarter}[/][/]")
    else:
        console.print("[yellow]⚠ Could not parse quarter from sheet name "
                      f"'{rate_sheet}' — Quarter_Year filter will be skipped[/]")

    q_for_name = (target_quarter or "AllQ").replace(" ", "_")
    out_name = f"FabSol_Compliance_{q_for_name}_{date.today():%Y-%m-%d}.xlsx"
    out_path = WORK_DIR / out_name

    with make_progress() as progress:
        df = load_tpm(tpm_path, tpm_sheet, progress)
        df = validate_and_filter(df, progress)
        df = add_helper_columns(df, target_quarter, progress)
        rc_index = build_rate_card_index(rate_path, rate_sheet, progress)
        df = evaluate(df, rc_index, progress)
        summary = build_summary(df)
        write_output(df, summary, out_path,
                     tpm_path, tpm_sheet, rate_path, rate_sheet,
                     target_quarter, progress)

    display_summary(df, out_path)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print(f"\n[yellow]Cancelled by user (v{__version__})[/]")
        sys.exit(0)
    except Exception as e:
        console.print(f"\n[red bold]❌ Error (v{__version__}): {e}[/]")
        console.print(Panel(traceback.format_exc(), title=f"Traceback — v{__version__}",
                            border_style="red", box=box.ROUNDED))
        sys.exit(1)