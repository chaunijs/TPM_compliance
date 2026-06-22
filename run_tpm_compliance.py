"""
TPM × FW Rate Card Compliance Validator — v1.6
==============================================
v1.6 changes:
  ✓ FW_Rate_Number is now a SET of values (slash-delimited string)
    extracted from multiple Excel columns per Period:
      - Weekly:        cols {AJ, L, V, X}
      - Bi/Tri-weekly: cols {J, N}
      - Monthly:       cols {H, AA}
  ✓ Each cell may contain '/' delimiter → split into the SET
  ✓ Numeric values → format as 1 decimal (e.g. "99.0")
  ✓ Non-numeric → keep as uppercase string (e.g. "BOGO", "2F1")
  ✓ Comply = Final_Rsp matches any numeric value in the SET (±0.5)
v1.5 features unchanged:
  ✓ Sales_Org == "7001" filter
  ✓ Quarter mismatch merged into Justification
  ✓ Drop unused columns (RSP CCBT, GAP, Unnamed, promo mechanic)
  ✓ Polars + EXACT MATCH Comply
"""

from __future__ import annotations
import re
import sys
import warnings
import logging
import contextlib
import io as _io
import traceback
from datetime import date
from pathlib import Path

import polars as pl
import pandas as pd          # used for messy rate-card layout only
import questionary
from questionary import Style as QStyle
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeElapsedColumn
from rich import box

console = Console()

warnings.filterwarnings("ignore", message=".*[Cc]ould not determine dtype.*")
warnings.filterwarnings("ignore", message=".*[Ff]alling back to string.*")
logging.getLogger("polars").setLevel(logging.ERROR)

# ============================================================
# CONFIG
# ============================================================
if getattr(sys, "frozen", False):
    WORK_DIR = Path(sys.executable).parent
else:
    WORK_DIR = Path(__file__).parent

OUT_NAME = f"TPM_FW_Compliance_Output_{date.today():%Y-%m-%d}.xlsx"
LOG_NAME = f"TPM_FW_Compliance_Log_{date.today():%Y-%m-%d}.txt"

DROP_COLUMNS = ["RSP CCBT", "GAP", "promo mechanic"]
VALID_SALES_ORG = "7001"

# ============================================================
# v1.6 NEW: Excel column SET mapping for FW_Rate_Number
# ============================================================
def col_letter_to_idx(letter: str) -> int:
    """Convert Excel column letter to 0-based index. A=0, Z=25, AA=26, AJ=35."""
    letter = letter.upper().strip()
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - ord('A') + 1)
    return n - 1

# Period → SET of Excel columns to pull from
PERIOD_COL_MAP = {
    "Weekly":     [col_letter_to_idx(c) for c in ("AJ", "L", "V", "X")],
    "Bi-weekly":  [col_letter_to_idx(c) for c in ("J", "N")],
    "Tri-weekly": [col_letter_to_idx(c) for c in ("J", "N")],
    "Monthly":    [col_letter_to_idx(c) for c in ("H", "AA")],
}

QSTYLE = QStyle([
    ("qmark",       "fg:#00aaff bold"),
    ("question",    "bold"),
    ("answer",      "fg:#00ff88 bold"),
    ("pointer",     "fg:#ff8800 bold"),
    ("highlighted", "fg:#00aaff bold"),
    ("selected",    "fg:#00ff88"),
    ("instruction", "fg:#888888 italic"),
])


# ============================================================
# UI HELPERS (unchanged from v1.5)
# ============================================================
def banner():
    console.print()
    console.print(Panel.fit(
        "[bold cyan]TPM × FW Rate Card Compliance Validator[/]\n"
        "[dim]v1.6 — Polars + SET-based FW_Rate_Number[/]\n"
        f"[dim]{WORK_DIR}[/]",
        border_style="cyan",
        box=box.DOUBLE,
    ))
    console.print()


def list_xlsx_files(folder: Path) -> list:
    files = sorted([
        p for p in folder.iterdir()
        if p.suffix.lower() in (".xlsx", ".xlsb", ".xls")
        and not p.name.startswith("~$")
        and not p.name.startswith("TPM_FW_Compliance_Output")
        and p.is_file()
    ], key=lambda p: p.stat().st_mtime, reverse=True)
    return files


def pick_file(files: list, prompt: str, suggest_keyword: str = "") -> Path:
    if suggest_keyword:
        kw = suggest_keyword.lower()
        matching = [f for f in files if kw in f.name.lower()]
        others   = [f for f in files if kw not in f.name.lower()]
        files    = matching + others

    choices = []
    for i, f in enumerate(files, 1):
        size_kb = f.stat().st_size / 1024
        mtime   = pd.Timestamp(f.stat().st_mtime, unit="s").strftime("%Y-%m-%d %H:%M")
        size_str = f"{size_kb/1024:6.1f} MB" if size_kb > 1024 else f"{size_kb:6.0f} KB"
        label = f"{i:>2}. {f.name:<55s}  {size_str}  {mtime}"
        if suggest_keyword and suggest_keyword.lower() in f.name.lower():
            label = "→ " + label
        else:
            label = "  " + label
        choices.append(questionary.Choice(label, value=f))

    answer = questionary.select(
        prompt,
        choices=choices,
        instruction="(↑↓ arrows + Enter, Ctrl-C to cancel)",
        style=QSTYLE,
        use_shortcuts=False,
    ).ask()
    if answer is None:
        console.print("[red]Cancelled.[/]")
        sys.exit(0)
    return answer


def pick_sheet(rate_path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(rate_path, read_only=True, data_only=False)
    sheets = wb.sheetnames
    wb.close()

    def sort_key(s):
        m = re.search(r"Q\s*([1-4])[\s,_-]*\s*(20\d{2})", s, re.IGNORECASE)
        if m:
            return (int(m.group(2)), int(m.group(1)))
        return (-1, -1)

    sheets_sorted = sorted(sheets, key=sort_key, reverse=True)

    choices = []
    for i, s in enumerate(sheets_sorted, 1):
        m = re.search(r"Q\s*([1-4])[\s,_-]*\s*(20\d{2})", s, re.IGNORECASE)
        tag = f"  [Q{m.group(1)} {m.group(2)}]" if m else ""
        prefix = "→ " if i == 1 else "  "
        choices.append(questionary.Choice(f"{prefix}{i:>2}. {s}{tag}", value=s))

    answer = questionary.select(
        f"📊 Which sheet from {rate_path.name}?",
        choices=choices,
        instruction="(↑↓ arrows + Enter, latest quarter pre-selected)",
        style=QSTYLE,
    ).ask()
    if answer is None:
        console.print("[red]Cancelled.[/]")
        sys.exit(0)
    return answer


# ============================================================
# POLARS PROCESSING
# ============================================================
def _columns_to_drop(df_columns: list) -> list:
    drop_lower = {c.lower() for c in DROP_COLUMNS}
    result = []
    for col in df_columns:
        c = str(col).strip()
        if c.lower() in drop_lower:
            result.append(col)
        elif c.lower().startswith("unnamed") or c.startswith("__UNNAMED__"):
            result.append(col)
    return result


def load_tpm(tpm_path: Path) -> pl.DataFrame:
    if tpm_path.suffix.lower() == ".xlsb":
        df_pd = pd.read_excel(tpm_path, sheet_name=0, engine="pyxlsb")
        df = pl.from_pandas(df_pd)
    else:
        with contextlib.redirect_stderr(_io.StringIO()):
            df = pl.read_excel(tpm_path, sheet_id=1, engine="calamine")

    initial_rows = df.height

    cols_to_drop = _columns_to_drop(df.columns)
    if cols_to_drop:
        df = df.drop(cols_to_drop)
        console.print(f"   [dim]Dropped {len(cols_to_drop)} unused column(s): "
                      f"{', '.join(cols_to_drop)}[/]")

    if "Sales_Org" in df.columns:
        df = df.with_columns(
            pl.col("Sales_Org").cast(pl.Utf8).str.strip_chars().alias("_sales_org_str")
        )
        before = df.height
        df = df.filter(pl.col("_sales_org_str") == VALID_SALES_ORG)
        df = df.drop("_sales_org_str")
        dropped = before - df.height
        if dropped > 0:
            console.print(f"   [dim]Filtered Sales_Org=='{VALID_SALES_ORG}': "
                          f"dropped {dropped:,} irrelevant rows ({before:,} → {df.height:,})[/]")
    else:
        console.print(f"   [yellow]⚠ Sales_Org column not found — skipping filter[/]")

    if df.height == 0:
        raise RuntimeError(
            f"After filtering Sales_Org=='{VALID_SALES_ORG}', no rows remain. "
            f"Original file had {initial_rows:,} rows."
        )

    return df


def add_helper_columns(df: pl.DataFrame) -> pl.DataFrame:
    required = ["PromoGroupProductDesc", "Instore_Start", "Instore_End",
                "TPM_InvestmentDescription", "RSP Promo"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Missing required columns: {missing}")

    df = df.with_columns(
        pl.col("TPM_InvestmentDescription").cast(pl.Utf8)
          .str.to_uppercase().alias("_an_upper")
    ).with_columns(
        pl.when(pl.col("_an_upper").str.contains("BOGO",   literal=True)).then(pl.lit("BOGO"))
        .when(pl.col("_an_upper").str.contains("2F1",    literal=True)).then(pl.lit("2F1"))
        .when(pl.col("_an_upper").str.contains("LAKSUE", literal=True)).then(pl.lit("LAKSUE"))
        .when(pl.col("_an_upper").str.contains("_PO",    literal=True)).then(pl.lit("PO"))
        .otherwise(pl.lit(""))
        .alias("Promo_Type")
    )

    df = df.with_columns(
        pl.when(pl.col("Promo_Type") == "LAKSUE")
            .then(pl.col("_an_upper").str.extract(r"_LAKSUE ?(\d+)", 1)
                                       .cast(pl.Float64, strict=False))
        .when(pl.col("Promo_Type").is_in(["PO", "2F1"]))
            .then(pl.col("_an_upper").str.extract(r"_PO ?(\d+)", 1)
                                       .cast(pl.Float64, strict=False))
        .otherwise(None)
        .alias("Promotion_Price")
    )

    df = df.with_columns([
        pl.col("Instore_Start").cast(pl.Datetime, strict=False),
        pl.col("Instore_End").cast(pl.Datetime,   strict=False),
    ])

    df = df.with_columns(
        (pl.col("Instore_End") - pl.col("Instore_Start"))
        .dt.total_days().alias("DateRange_Days")
    )

    df = df.with_columns(
        pl.when(pl.col("DateRange_Days").is_null()).then(pl.lit(""))
        .when(pl.col("DateRange_Days") <= 10.5).then(pl.lit("Weekly"))
        .when(pl.col("DateRange_Days") <= 17.5).then(pl.lit("Bi-weekly"))
        .when(pl.col("DateRange_Days") <= 25.5).then(pl.lit("Tri-weekly"))
        .otherwise(pl.lit("Monthly"))
        .alias("Period")
    )

    df = df.with_columns(
        pl.when((pl.col("RSP Promo").is_not_null()) & (pl.col("RSP Promo") > 0))
        .then(pl.col("RSP Promo"))
        .otherwise(pl.col("Promotion_Price"))
        .alias("Final_Rsp")
    )

    df = df.with_columns(
        pl.when(pl.col("Instore_Start").is_null()).then(None)
        .otherwise(
            pl.format("Q{}_{}",
                ((pl.col("Instore_Start").dt.month() - 1) // 3 + 1),
                pl.col("Instore_Start").dt.year()
            )
        ).alias("Quarter_Year")
    )

    return df.drop("_an_upper")


# ============================================================
# v1.6 NEW: Parse a cell that may contain '/' delimiter into a SET
# ============================================================
def parse_set_cell(v) -> list:
    """
    Parse a rate-card cell into a list of values.
    - NaN/empty   → []
    - Numeric     → [round(value, 1)]   (1 decimal)
    - Text with '/' → split, then numeric→float, else uppercase string
    - Text without '/' → numeric→float, else uppercase string
    """
    if pd.isna(v):
        return []
    if isinstance(v, (int, float)):
        return [round(float(v), 1)]
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return []
    parts = [p.strip() for p in s.split("/") if p.strip()]
    out = []
    for p in parts:
        try:
            out.append(round(float(p), 1))
        except ValueError:
            # Strip a trailing % if any, retry once
            try:
                out.append(round(float(p.rstrip("%").strip()), 1))
            except ValueError:
                out.append(p.upper())
    return out


def format_set(values: list) -> str | None:
    """Format a SET list back into display string '99.0/109.0/BOGO'."""
    if not values:
        return None
    parts = []
    for v in values:
        if isinstance(v, str):
            parts.append(v)
        else:
            parts.append(f"{v:.1f}")
    return "/".join(parts)


# ============================================================
# v1.6 UPDATED: Build rate lookup using Excel-column SETS
# ============================================================
def build_rate_lookup(rate_path: Path, sheet_name: str) -> tuple:
    """
    Read rate-card sheet → return (lookup_df, quarter_tag).
    Builds SET of values from PERIOD_COL_MAP for each (PPG, Period).
    """
    df_raw = pd.read_excel(rate_path, sheet_name=sheet_name, header=None)

    m = re.search(r"Q\s*([1-4])[\s,_-]*\s*(20\d{2})", sheet_name, re.IGNORECASE)
    quarter = (f"Q{m.group(1)}_{m.group(2)}" if m
               else f"Q{(date.today().month - 1)//3 + 1}_{date.today().year}")

    hdr_row = df_raw.apply(
        lambda r: r.astype(str).str.contains("Promo Group", case=False, na=False).any(),
        axis=1).idxmax()

    def locate(contains):
        row = df_raw.iloc[hdr_row].astype(str)
        hits = [i for i, v in row.items() if contains.lower() in v.lower()]
        return hits[0] if hits else None

    ppg_col = locate("Promo Group")
    if ppg_col is None:
        raise RuntimeError(f"Could not find 'Promo Group' column in '{sheet_name}'")

    max_col_needed = max(max(cols) for cols in PERIOD_COL_MAP.values())
    if df_raw.shape[1] <= max_col_needed:
        console.print(f"   [yellow]⚠ Sheet has only {df_raw.shape[1]} columns, "
                      f"but logic expects up to col index {max_col_needed} (column "
                      f"{chr(ord('A')+max_col_needed) if max_col_needed < 26 else '...'}).[/]")

    records = []
    for _, row in df_raw.iloc[hdr_row + 1:].iterrows():
        ppg = row.iloc[ppg_col]
        if pd.isna(ppg): continue
        ppg_str = str(ppg).strip()
        if not ppg_str or ppg_str.lower() in ("nan", "note"): continue

        for period, col_indices in PERIOD_COL_MAP.items():
            value_set = []
            for col_idx in col_indices:
                if col_idx < df_raw.shape[1]:
                    parsed = parse_set_cell(row.iloc[col_idx])
                    value_set.extend(parsed)

            # Dedupe preserving order (treat float and string as distinct keys)
            seen = set()
            unique_set = []
            for v in value_set:
                key = ("S", v) if isinstance(v, str) else ("N", round(float(v), 1))
                if key not in seen:
                    seen.add(key)
                    unique_set.append(v)

            if unique_set:
                records.append({
                    "PPG":            ppg_str,
                    "Period":         period,
                    "FW_Rate_Number": format_set(unique_set),
                })

    if not records:
        raise RuntimeError(f"No usable rates from '{sheet_name}'")

    lookup = pl.DataFrame(records)
    return lookup, quarter


def attach_rate(tpm: pl.DataFrame, lookup: pl.DataFrame, quarter: str) -> pl.DataFrame:
    """Left-join lookup via (PPG, Period)."""
    tpm = tpm.with_columns(
        pl.col("PromoGroupProductDesc").cast(pl.Utf8)
          .str.strip_chars().alias("_ppg_key")
    )
    tpm = tpm.join(
        lookup.rename({"PPG": "_ppg_key"}),
        on=["_ppg_key", "Period"],
        how="left"
    ).drop("_ppg_key")
    tpm = tpm.with_columns(pl.lit(quarter).alias("Rate_Card_Source"))
    return tpm


# ============================================================
# v1.6 UPDATED: Justification checks Final_Rsp ∈ FW_Rate_Number SET
# ============================================================
def _final_in_set(rate_str: str | None, final_rsp: float | None) -> bool | None:
    """
    Return True if Final_Rsp matches any NUMERIC value in the '/' set
    (tolerance ±0.5). Strings in the set (e.g. 'BOGO') don't match numbers.
    Return None if inputs are null (handled by outer when/then).
    """
    if rate_str is None or final_rsp is None:
        return None
    if not rate_str:
        return False
    for part in str(rate_str).split("/"):
        part = part.strip()
        try:
            if abs(float(part) - float(final_rsp)) <= 0.5:
                return True
        except ValueError:
            continue
    return False


def add_justification(tpm: pl.DataFrame) -> pl.DataFrame:
    """
    Priority:
      1. NO rate card available  (FW_Rate_Number null/empty)
      2. Missing Final_Rsp
      3. ⚠ Quarter mismatch
      4. Comply        (Final_Rsp ∈ FW_Rate_Number SET, ±0.5)
      5. NOT Comply
    """
    # Pre-compute set-membership boolean using map_elements
    tpm = tpm.with_columns(
        pl.struct(["FW_Rate_Number", "Final_Rsp"])
          .map_elements(
              lambda s: _final_in_set(s["FW_Rate_Number"], s["Final_Rsp"]),
              return_dtype=pl.Boolean
          ).alias("_in_set")
    )

    tpm = tpm.with_columns(
        pl.when(pl.col("FW_Rate_Number").is_null() | (pl.col("FW_Rate_Number") == ""))
            .then(pl.lit("NO rate card available"))
        .when(pl.col("Final_Rsp").is_null())
            .then(pl.lit("Missing Final_Rsp"))
        .when(
            pl.col("Quarter_Year").is_not_null()
            & pl.col("Rate_Card_Source").is_not_null()
            & (pl.col("Quarter_Year") != pl.col("Rate_Card_Source"))
        ).then(pl.lit("⚠ Quarter mismatch"))
        .when(pl.col("_in_set") == True)
            .then(pl.lit("Comply"))
        .otherwise(pl.lit("NOT Comply"))
        .alias("Justification")
    ).drop("_in_set")

    return tpm


def build_summary(tpm: pl.DataFrame) -> pl.DataFrame:
    grp = [c for c in ["Customer", "Brand", "Period"] if c in tpm.columns]
    if not grp:
        return pl.DataFrame({"Note": ["No grouping columns found"]})

    summary = tpm.group_by(grp).agg([
        (pl.col("Justification") == "Comply").sum().alias("Comply"),
        (pl.col("Justification") == "NOT Comply").sum().alias("NOT Comply"),
        (pl.col("Justification") == "⚠ Quarter mismatch").sum().alias("⚠ Quarter mismatch"),
        (pl.col("Justification") == "NO rate card available").sum().alias("NO rate card available"),
        (pl.col("Justification") == "Missing Final_Rsp").sum().alias("Missing Final_Rsp"),
    ]).with_columns(
        (pl.col("Comply") + pl.col("NOT Comply") + pl.col("⚠ Quarter mismatch") +
         pl.col("NO rate card available") + pl.col("Missing Final_Rsp")).alias("Total")
    ).with_columns(
        (pl.col("Comply") / pl.col("Total") * 100).round(1).alias("Comply_%")
    ).sort("Total", descending=True)

    return summary


# ============================================================
# OUTPUT (unchanged from v1.5)
# ============================================================
def save_output(tpm: pl.DataFrame, summary: pl.DataFrame, out_path: Path):
    final_drop = _columns_to_drop(tpm.columns)
    if final_drop:
        tpm = tpm.drop(final_drop)

    if out_path.exists():
        try:
            out_path.unlink()
        except PermissionError:
            out_path = out_path.with_stem(out_path.stem + "_v2")
            console.print(f"   [yellow]⚠ Previous output locked → writing to {out_path.name}[/]")

    float_fmt = {"num_format": "0.0"}
    int_fmt   = {"num_format": "0"}
    date_fmt  = {"num_format": "yyyy-mm-dd"}
    col_formats = {}
    for col, dtype in zip(tpm.columns, tpm.dtypes):
        if dtype in (pl.Float32, pl.Float64):
            col_formats[col] = float_fmt
        elif dtype in (pl.Int8, pl.Int16, pl.Int32, pl.Int64,
                       pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64):
            col_formats[col] = int_fmt
        elif dtype in (pl.Date, pl.Datetime):
            col_formats[col] = date_fmt

    import xlsxwriter
    with xlsxwriter.Workbook(out_path) as wb:
        tpm.write_excel(
            workbook=wb,
            worksheet="Compliance",
            autofit=True,
            freeze_panes=(1, 0),
            autofilter=True,
            header_format={
                "bold": True, "bg_color": "#305496",
                "font_color": "white", "align": "center", "valign": "vcenter",
            },
            column_formats=col_formats,
            conditional_formats={
                "Justification": [
                    {"type": "cell", "criteria": "==", "value": '"Comply"',
                     "format": {"bg_color": "#C6EFCE", "font_color": "#006100"}},
                    {"type": "cell", "criteria": "==", "value": '"NOT Comply"',
                     "format": {"bg_color": "#FFC7CE", "font_color": "#9C0006"}},
                    {"type": "cell", "criteria": "==", "value": '"⚠ Quarter mismatch"',
                     "format": {"bg_color": "#FCE4D6", "font_color": "#974706", "bold": True}},
                    {"type": "cell", "criteria": "==", "value": '"NO rate card available"',
                     "format": {"bg_color": "#FFEB9C", "font_color": "#9C5700"}},
                    {"type": "cell", "criteria": "==", "value": '"Missing Final_Rsp"',
                     "format": {"bg_color": "#D9D9D9", "font_color": "#555555"}},
                ],
            },
        )

        summary_col_formats = {}
        for col, dtype in zip(summary.columns, summary.dtypes):
            if dtype in (pl.Float32, pl.Float64):
                summary_col_formats[col] = float_fmt
            elif dtype in (pl.Int8, pl.Int16, pl.Int32, pl.Int64,
                           pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64):
                summary_col_formats[col] = int_fmt

        summary.write_excel(
            workbook=wb,
            worksheet="Summary",
            autofit=True,
            freeze_panes=(1, 0),
            autofilter=True,
            header_format={
                "bold": True, "bg_color": "#305496",
                "font_color": "white", "align": "center",
            },
            column_formats=summary_col_formats,
            conditional_formats={
                "Comply_%": [
                    {"type": "3_color_scale",
                     "min_color": "#F8696B", "mid_color": "#FFEB84", "max_color": "#63BE7B"}
                ]
            },
        )

    if not out_path.exists() or out_path.stat().st_size < 1024:
        raise RuntimeError(f"Output file empty/corrupt: {out_path}")
    chk = pd.read_excel(out_path, sheet_name="Compliance", nrows=5)
    if chk.empty:
        raise RuntimeError("Saved file failed verification.")
    return out_path


# ============================================================
# RICH SUMMARY DISPLAY (unchanged from v1.5)
# ============================================================
def display_summary(tpm: pl.DataFrame, out_path: Path):
    total = tpm.height
    counts = tpm["Justification"].value_counts().sort("count", descending=True)

    table = Table(title="📊 Justification Breakdown",
                  box=box.ROUNDED, header_style="bold cyan")
    table.add_column("Category", style="bold")
    table.add_column("Count", justify="right")
    table.add_column("%",     justify="right")
    table.add_column("Bar",   justify="left")

    color_map = {
        "Comply":                  "green",
        "NOT Comply":              "red",
        "⚠ Quarter mismatch":      "dark_orange",
        "NO rate card available":  "yellow",
        "Missing Final_Rsp":       "white",
    }
    for row in counts.iter_rows(named=True):
        cat = row["Justification"]
        n   = row["count"]
        pct = n / total * 100
        bar_width = int(pct / 2)
        bar = "█" * bar_width
        color = color_map.get(cat, "white")
        table.add_row(
            f"[{color}]{cat}[/]",
            f"{n:,}",
            f"{pct:5.1f}%",
            f"[{color}]{bar}[/]"
        )
    table.add_section()
    table.add_row("[bold]Total[/]", f"[bold]{total:,}[/]", "100.0%", "")
    console.print(table)

    size_mb = out_path.stat().st_size / 1024 / 1024
    console.print(Panel.fit(
        f"[green]✅ Output saved successfully![/]\n\n"
        f"📄 [bold]{out_path.name}[/]\n"
        f"📁 [dim]{out_path.parent}[/]\n"
        f"💾 [dim]{size_mb:.1f} MB · {total:,} rows[/]\n\n"
        f"[dim]Sheets:[/]\n"
        f"  • [cyan]Compliance[/] — all rows, color-coded\n"
        f"  • [cyan]Summary[/]     — pivot by Customer × Brand × Period",
        border_style="green",
        box=box.ROUNDED,
    ))


# ============================================================
# MAIN
# ============================================================
def main():
    banner()

    files = list_xlsx_files(WORK_DIR)
    if not files:
        console.print(f"[red]❌ No xlsx files found in {WORK_DIR}[/]")
        sys.exit(1)

    console.print(f"[dim]Found {len(files)} Excel file(s) in working folder[/]\n")

    tpm_path = pick_file(files, "📂 Which file is the TPM file?",
                          suggest_keyword="tpm")
    console.print(f"   [green]✓[/] TPM file: [bold]{tpm_path.name}[/]\n")

    remaining = [f for f in files if f != tpm_path]
    if not remaining:
        console.print("[red]❌ Only one file in folder. Need a separate FW rate card file.[/]")
        sys.exit(1)
    rate_path = pick_file(remaining, "📂 Which file is the FW Rate Card file?",
                           suggest_keyword="fw rate card")
    console.print(f"   [green]✓[/] Rate card: [bold]{rate_path.name}[/]\n")

    sheet_name = pick_sheet(rate_path)
    console.print(f"   [green]✓[/] Sheet: [bold]{sheet_name}[/]\n")

    with Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:

        t1 = progress.add_task("Loading TPM with Polars...", total=6)

        tpm = load_tpm(tpm_path)
        progress.console.print(f"   [dim]Loaded {tpm.height:,} rows × {tpm.width} cols[/]")
        progress.advance(t1)

        progress.update(t1, description="Deriving helper columns (vectorized)...")
        tpm = add_helper_columns(tpm)
        progress.advance(t1)

        progress.update(t1, description="Building rate-card SET lookup (v1.6)...")
        lookup, quarter = build_rate_lookup(rate_path, sheet_name)
        progress.console.print(f"   [dim]Lookup table: {lookup.height:,} (PPG × Period) entries "
                                f"from {quarter}[/]")
        progress.advance(t1)

        progress.update(t1, description="Joining rate card SET to TPM...")
        tpm = attach_rate(tpm, lookup, quarter)
        progress.advance(t1)

        progress.update(t1, description="Computing Justification (SET membership)...")
        tpm = add_justification(tpm)
        progress.advance(t1)

        progress.update(t1, description="Building summary & saving xlsx...")
        summary = build_summary(tpm)
        out_path = WORK_DIR / OUT_NAME
        out_path = save_output(tpm, summary, out_path)
        progress.advance(t1)

    console.print()
    display_summary(tpm, out_path)

    console.print("\n[bold]Self-tests:[/]")
    no_rc = (tpm["Justification"] == "NO rate card available").mean()
    icon = "[green]✅[/]" if no_rc < 0.95 else "[red]❌[/]"
    console.print(f"  {icon} NO-rate-card share: {no_rc:.1%}")

    bogo = tpm.filter(pl.col("Promo_Type") == "BOGO")
    if bogo.height:
        ok = bogo["Promotion_Price"].is_null().all()
        icon = "[green]✅[/]" if ok else "[red]❌[/]"
        console.print(f"  {icon} BOGO blank-price: {ok}")

    for pt in ["PO", "2F1", "LAKSUE"]:
        sub = tpm.filter(pl.col("Promo_Type") == pt)
        if sub.height:
            rate = sub["Promotion_Price"].is_not_null().mean()
            icon = "[green]✅[/]" if rate > 0.8 else "[yellow]⚠️[/]"
            console.print(f"  {icon} {pt:<6s} extraction: {rate:.0%} of {sub.height:,} rows")

    # Show a few sample FW_Rate_Number values to verify SET parsing
    if tpm.height > 0 and "FW_Rate_Number" in tpm.columns:
        sample = (tpm.filter(pl.col("FW_Rate_Number").is_not_null())
                     .select(["Period", "FW_Rate_Number"]).unique()
                     .head(5))
        if sample.height > 0:
            console.print("\n[bold]Sample FW_Rate_Number SETs (v1.6):[/]")
            for row in sample.iter_rows(named=True):
                console.print(f"  [dim]{row['Period']:<12s}[/] → [cyan]{row['FW_Rate_Number']}[/]")

    console.print("\n[bold green]🎉 Done![/]\n")


# ============================================================
if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[yellow]Cancelled by user[/]")
        sys.exit(0)
    except Exception as e:
        console.print(f"\n[red bold]❌ Error: {e}[/]")
        console.print(Panel(traceback.format_exc(), title="Traceback",
                             border_style="red", box=box.ROUNDED))
        sys.exit(1)